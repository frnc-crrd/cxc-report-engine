"""Tests para src/excel_formatter.py.

No requiere Firebird ni PostgreSQL.
"""

from __future__ import annotations

from pathlib import Path
from unittest.mock import MagicMock, patch

import pandas as pd
import pytest
from openpyxl import load_workbook

from src.excel_formatter import escribir_hoja, exportar_excel, extraer_banda


# ---- exportar_excel ----------------------------------------------------------

def test_exportar_excel_creates_file(tmp_path: Path) -> None:
    df = pd.DataFrame({"NOMBRE_CLIENTE": ["A", "B"], "SALDO_FACTURA": [100.0, 200.0]})
    path = exportar_excel({"hoja1": df}, "reporte_test", output_dir=tmp_path, timestamp="ts")
    assert path.exists()
    assert path.suffix == ".xlsx"


def test_exportar_excel_filename_includes_timestamp(tmp_path: Path) -> None:
    df = pd.DataFrame({"COL": [1]})
    path = exportar_excel({"h": df}, "base", output_dir=tmp_path, timestamp="20260101")
    assert "20260101" in path.name


def test_exportar_excel_without_timestamp(tmp_path: Path) -> None:
    df = pd.DataFrame({"COL": [1]})
    path = exportar_excel({"h": df}, "base", output_dir=tmp_path)
    assert path.name == "BASE.xlsx"


def test_exportar_excel_empty_dict_writes_contingencia(tmp_path: Path) -> None:
    path = exportar_excel({}, "vacio", output_dir=tmp_path)
    wb = load_workbook(path)
    assert "Sin Datos" in wb.sheetnames


def test_exportar_excel_none_df_skipped(tmp_path: Path) -> None:
    df = pd.DataFrame({"COL": [1]})
    path = exportar_excel({"real": df, "nulo": None}, "test", output_dir=tmp_path)
    wb = load_workbook(path)
    assert "real" in wb.sheetnames
    assert "nulo" not in wb.sheetnames


def test_exportar_excel_empty_df_skipped(tmp_path: Path) -> None:
    df = pd.DataFrame({"COL": [1]})
    vacio = pd.DataFrame()
    path = exportar_excel({"real": df, "vacio": vacio}, "test", output_dir=tmp_path)
    wb = load_workbook(path)
    assert "real" in wb.sheetnames
    assert "vacio" not in wb.sheetnames


def test_exportar_excel_respects_sheet_order(tmp_path: Path) -> None:
    dfs = {
        "primero": pd.DataFrame({"A": [1]}),
        "segundo": pd.DataFrame({"B": [2]}),
        "tercero": pd.DataFrame({"C": [3]}),
    }
    path = exportar_excel(
        dfs, "orden", output_dir=tmp_path,
        orden_hojas=["tercero", "primero", "segundo"],
    )
    wb = load_workbook(path)
    assert wb.sheetnames == ["tercero", "primero", "segundo"]


def test_exportar_excel_creates_parent_dir(tmp_path: Path) -> None:
    subdir = tmp_path / "nuevo" / "subdir"
    df = pd.DataFrame({"X": [1]})
    path = exportar_excel({"h": df}, "test", output_dir=subdir)
    assert path.exists()


def test_exportar_excel_sheet_name_truncated_to_31(tmp_path: Path) -> None:
    nombre_largo = "a" * 40
    df = pd.DataFrame({"COL": [1]})
    path = exportar_excel({nombre_largo: df}, "test", output_dir=tmp_path)
    wb = load_workbook(path)
    assert all(len(s) <= 31 for s in wb.sheetnames)


# ---- extraer_banda -----------------------------------------------------------

def test_extraer_banda_con_columna(tmp_path: Path) -> None:
    df = pd.DataFrame({"A": [1, 2], "_BAND_GROUP": [0, 1]})
    df_limpio, band = extraer_banda(df)
    assert "_BAND_GROUP" not in df_limpio.columns
    assert list(band) == [0, 1]


def test_extraer_banda_sin_columna(tmp_path: Path) -> None:
    df = pd.DataFrame({"A": [1, 2]})
    df_limpio, band = extraer_banda(df)
    assert band is None
    assert list(df_limpio.columns) == ["A"]


def test_extraer_banda_no_modifica_original() -> None:
    df = pd.DataFrame({"A": [1], "_BAND_GROUP": [0]})
    extraer_banda(df)
    assert "_BAND_GROUP" in df.columns


# ---- seguridad: proteccion de hojas con password ----------------------------

def test_escribir_hoja_protegida_password_none_lanza_error() -> None:
    """TDD S1-A: escribir_hoja con protegida=True y password=None debe lanzar ValueError."""
    df = pd.DataFrame({"COL": [1]})
    mock_writer = MagicMock()
    with pytest.raises(ValueError, match="EXCEL_SHEET_PASSWORD"):
        escribir_hoja(mock_writer, "hoja", df, protegida=True, password=None)


def test_escribir_hoja_protegida_password_vacio_lanza_error() -> None:
    """TDD S1-A: escribir_hoja con protegida=True y password vacio debe lanzar ValueError."""
    df = pd.DataFrame({"COL": [1]})
    mock_writer = MagicMock()
    with pytest.raises(ValueError, match="EXCEL_SHEET_PASSWORD"):
        escribir_hoja(mock_writer, "hoja", df, protegida=True, password="")


def test_exportar_excel_hoja_protegida_sin_password_lanza_error(tmp_path: Path) -> None:
    """TDD S1-A: exportar una hoja protegida cuando EXCEL_SHEET_PASSWORD no esta definida.

    Patch explicito de SHEET_PASSWORDS para garantizar que el test sea determinista
    independientemente de si la variable de entorno esta definida en el entorno CI.
    """
    df = pd.DataFrame({"COL": [1]})
    with patch("src.excel_formatter.SHEET_PASSWORDS", {"registros_totales_cxc": None}):
        with pytest.raises(ValueError, match="EXCEL_SHEET_PASSWORD"):
            exportar_excel({"registros_totales_cxc": df}, "test", output_dir=tmp_path)


def test_exportar_excel_hoja_protegida_con_password_activa_proteccion(tmp_path: Path) -> None:
    """TDD S1-A: una hoja protegida con password valido debe tener protection.sheet == True."""
    df = pd.DataFrame({"COL": [1]})
    with patch("src.excel_formatter.SHEET_PASSWORDS", {"registros_totales_cxc": "password_test_seguro"}):
        path = exportar_excel({"registros_totales_cxc": df}, "test", output_dir=tmp_path)
    wb = load_workbook(path)
    assert wb["registros_totales_cxc"].protection.sheet is True


def test_escribir_hoja_no_protegida_no_requiere_password(tmp_path: Path) -> None:
    """Regresion S1-A: una hoja no protegida no debe exigir password."""
    df = pd.DataFrame({"COL": [1]})
    filepath = tmp_path / "test.xlsx"
    with pd.ExcelWriter(filepath, engine="openpyxl") as writer:
        escribir_hoja(writer, "hoja_normal", df, protegida=False)
    wb = load_workbook(filepath)
    assert wb["hoja_normal"].protection.sheet is False


# ---- S2-C: formato por columna (column_dimensions) --------------------------

def test_exportar_excel_columna_moneda_tiene_formato_en_column_dimensions(tmp_path: Path) -> None:
    """TDD S2-C: columnas de moneda deben usar column_dimensions.number_format.

    Con el nuevo enfoque (column_dimensions), el formato se aplica a nivel de
    columna en lugar de celda a celda. El column_dimensions de una columna
    SALDO_FACTURA debe ser '#,##0.00'.
    """
    from openpyxl.utils import get_column_letter
    df = pd.DataFrame({
        "NOMBRE_CLIENTE": ["EMPRESA A", "EMPRESA B"],
        "SALDO_FACTURA":  [1160.0,     580.0],
    })
    path = exportar_excel({"hoja": df}, "test_fmt", output_dir=tmp_path)
    wb = load_workbook(path)
    ws = wb["hoja"]
    saldo_col = list(df.columns).index("SALDO_FACTURA") + 1
    letter = get_column_letter(saldo_col)
    assert ws.column_dimensions[letter].number_format == "#,##0.00", (
        f"column_dimensions['{letter}'].number_format debe ser '#,##0.00'"
    )


def test_exportar_excel_columna_fecha_solo_tiene_formato_yyyy_mm_dd(tmp_path: Path) -> None:
    """Fechas de emision y vencimiento deben usar YYYY/MM/DD sin hora."""
    from openpyxl.utils import get_column_letter
    df = pd.DataFrame({
        "NOMBRE_CLIENTE":    ["EMPRESA A"],
        "FECHA_EMISION":     pd.to_datetime(["2024-01-15"]),
        "FECHA_VENCIMIENTO": pd.to_datetime(["2024-02-15"]),
    })
    path = exportar_excel({"hoja": df}, "test_fecha", output_dir=tmp_path)
    wb = load_workbook(path)
    ws = wb["hoja"]
    for col_name in ["FECHA_EMISION", "FECHA_VENCIMIENTO"]:
        col_idx = list(df.columns).index(col_name) + 1
        letter = get_column_letter(col_idx)
        assert ws.column_dimensions[letter].number_format == "YYYY/MM/DD", (
            f"column_dimensions para {col_name} debe ser 'YYYY/MM/DD'"
        )


def test_exportar_excel_columna_fecha_hora_tiene_formato_con_tiempo(tmp_path: Path) -> None:
    """Columnas de auditoria (FECHA_HORA_*) deben preservar la hora."""
    from openpyxl.utils import get_column_letter
    df = pd.DataFrame({
        "NOMBRE_CLIENTE":         ["EMPRESA A"],
        "FECHA_HORA_CREACION":    pd.to_datetime(["2024-01-15 08:30:00"]),
        "FECHA_HORA_ULT_MODIF":   pd.to_datetime(["2024-01-16 12:00:00"]),
        "FECHA_HORA_CANCELACION": pd.to_datetime(["2024-01-17 15:45:00"]),
    })
    path = exportar_excel({"hoja": df}, "test_fecha_hora", output_dir=tmp_path)
    wb = load_workbook(path)
    ws = wb["hoja"]
    for col_name in ["FECHA_HORA_CREACION", "FECHA_HORA_ULT_MODIF", "FECHA_HORA_CANCELACION"]:
        col_idx = list(df.columns).index(col_name) + 1
        letter = get_column_letter(col_idx)
        assert ws.column_dimensions[letter].number_format == "YYYY/MM/DD HH:MM:SS", (
            f"column_dimensions para {col_name} debe ser 'YYYY/MM/DD HH:MM:SS'"
        )


def test_exportar_excel_columna_entero_usa_formato_sin_separador(tmp_path: Path) -> None:
    """Columnas enteras (NUM_*, DIAS_*) no deben usar separador de miles."""
    from openpyxl.utils import get_column_letter
    df = pd.DataFrame({
        "NOMBRE_CLIENTE":        ["EMPRESA A"],
        "NUM_FACTURAS_TOTALES":  [42],
        "DIAS_VENCIDO_MAX":      [30],
    })
    path = exportar_excel({"hoja": df}, "test_entero", output_dir=tmp_path)
    wb = load_workbook(path)
    ws = wb["hoja"]
    for col_name in ["NUM_FACTURAS_TOTALES", "DIAS_VENCIDO_MAX"]:
        col_idx = list(df.columns).index(col_name) + 1
        letter = get_column_letter(col_idx)
        assert ws.column_dimensions[letter].number_format == "0", (
            f"column_dimensions para {col_name} debe ser '0'"
        )


def test_exportar_excel_filas_tienen_altura_fija(tmp_path: Path) -> None:
    """Todas las filas deben tener altura fija para que no se expandan con el contenido."""
    df = pd.DataFrame({"NOMBRE_CLIENTE": ["A", "B"], "SALDO_FACTURA": [100.0, 200.0]})
    path = exportar_excel({"hoja": df}, "test_height", output_dir=tmp_path)
    wb = load_workbook(path)
    ws = wb["hoja"]
    for row_idx in range(1, len(df) + 2):
        height = ws.row_dimensions[row_idx].height
        assert height is not None and height > 0, (
            f"Fila {row_idx} debe tener altura fija definida explicitamente"
        )


def test_exportar_excel_columnas_numericas_alineadas_a_la_derecha(tmp_path: Path) -> None:
    """Celdas de columnas de moneda deben tener alineacion horizontal 'right'."""
    df = pd.DataFrame({"NOMBRE_CLIENTE": ["EMPRESA A"], "SALDO_FACTURA": [1000.0]})
    path = exportar_excel({"hoja": df}, "test_align", output_dir=tmp_path)
    wb = load_workbook(path)
    ws = wb["hoja"]
    saldo_col = list(df.columns).index("SALDO_FACTURA") + 1
    celda_dato = ws.cell(row=2, column=saldo_col)
    assert celda_dato.alignment.horizontal == "right", (
        "Las celdas de SALDO_FACTURA deben estar alineadas a la derecha"
    )


def test_exportar_excel_encabezados_centrados(tmp_path: Path) -> None:
    """Todos los encabezados deben estar centrados horizontalmente."""
    df = pd.DataFrame({"NOMBRE_CLIENTE": ["A"], "SALDO_FACTURA": [1.0], "FECHA_EMISION": pd.to_datetime(["2024-01-01"])})
    path = exportar_excel({"hoja": df}, "test_header_center", output_dir=tmp_path)
    wb = load_workbook(path)
    ws = wb["hoja"]
    for col_idx in range(1, len(df.columns) + 1):
        header = ws.cell(row=1, column=col_idx)
        assert header.alignment.horizontal == "center", (
            f"Encabezado de columna {col_idx} debe estar centrado"
        )


# ---- Formatos a nivel de CELDA (no solo column_dimensions) ------------------

def test_celda_moneda_tiene_number_format_miles_en_celda(tmp_path: Path) -> None:
    """El number_format de la celda de datos debe ser '#,##0.00', no solo el de la columna."""
    from openpyxl.utils import get_column_letter
    df = pd.DataFrame({"NOMBRE_CLIENTE": ["A"], "SALDO_FACTURA": [1234.56]})
    path = exportar_excel({"hoja": df}, "test_fmt_celda", output_dir=tmp_path)
    wb = load_workbook(path)
    ws = wb["hoja"]
    col_idx = list(df.columns).index("SALDO_FACTURA") + 1
    assert ws.cell(row=2, column=col_idx).number_format == "#,##0.00", (
        "La celda de SALDO_FACTURA debe tener number_format '#,##0.00' a nivel de celda"
    )


def test_celda_entero_tiene_number_format_0_en_celda(tmp_path: Path) -> None:
    """El number_format de celda de entero debe ser '0' (sin separador de miles)."""
    df = pd.DataFrame({"NOMBRE_CLIENTE": ["A"], "NUM_FACTURAS_TOTALES": [5]})
    path = exportar_excel({"hoja": df}, "test_entero_celda", output_dir=tmp_path)
    wb = load_workbook(path)
    ws = wb["hoja"]
    col_idx = list(df.columns).index("NUM_FACTURAS_TOTALES") + 1
    assert ws.cell(row=2, column=col_idx).number_format == "0", (
        "La celda de NUM_FACTURAS_TOTALES debe tener number_format '0' a nivel de celda"
    )


def test_celda_fecha_emision_tiene_number_format_yyyy_mm_dd_en_celda(tmp_path: Path) -> None:
    """El number_format de celda de FECHA_EMISION debe ser 'YYYY/MM/DD'."""
    df = pd.DataFrame({"NOMBRE_CLIENTE": ["A"], "FECHA_EMISION": pd.to_datetime(["2024-01-15"])})
    path = exportar_excel({"hoja": df}, "test_fecha_celda", output_dir=tmp_path)
    wb = load_workbook(path)
    ws = wb["hoja"]
    col_idx = list(df.columns).index("FECHA_EMISION") + 1
    assert ws.cell(row=2, column=col_idx).number_format == "YYYY/MM/DD", (
        "La celda de FECHA_EMISION debe tener number_format 'YYYY/MM/DD' a nivel de celda"
    )


def test_celda_fecha_hora_tiene_number_format_con_hora_en_celda(tmp_path: Path) -> None:
    """El number_format de celda de FECHA_HORA_ULT_MODIF debe incluir HH:MM:SS."""
    df = pd.DataFrame({
        "NOMBRE_CLIENTE":       ["A"],
        "FECHA_HORA_ULT_MODIF": pd.to_datetime(["2024-01-15 09:30:00"]),
    })
    path = exportar_excel({"hoja": df}, "test_fecha_hora_celda", output_dir=tmp_path)
    wb = load_workbook(path)
    ws = wb["hoja"]
    col_idx = list(df.columns).index("FECHA_HORA_ULT_MODIF") + 1
    assert ws.cell(row=2, column=col_idx).number_format == "YYYY/MM/DD HH:MM:SS", (
        "La celda de FECHA_HORA_ULT_MODIF debe tener number_format con hora a nivel de celda"
    )


def test_celda_porcentaje_tiene_number_format_pct_en_celda(tmp_path: Path) -> None:
    """El number_format de celda de PCT_VENCIDO debe ser '0.00%'."""
    df = pd.DataFrame({"NOMBRE_CLIENTE": ["A"], "PCT_VENCIDO": [0.25]})
    path = exportar_excel({"hoja": df}, "test_pct_celda", output_dir=tmp_path)
    wb = load_workbook(path)
    ws = wb["hoja"]
    col_idx = list(df.columns).index("PCT_VENCIDO") + 1
    assert ws.cell(row=2, column=col_idx).number_format == "0.00%", (
        "La celda de PCT_VENCIDO debe tener number_format '0.00%' a nivel de celda"
    )


def test_exportar_excel_encabezado_no_hereda_formato_de_columna(tmp_path: Path) -> None:
    """TDD S2-C: la celda de encabezado no debe heredar el formato numerico de la columna.

    Si la celda de encabezado no tiene formato explicito '@', el formato de
    column_dimensions se aplica a la cabecera, lo que es incorrecto visualmente.
    """
    df = pd.DataFrame({"SALDO_FACTURA": [100.0]})
    path = exportar_excel({"hoja": df}, "test_header_fmt", output_dir=tmp_path)
    wb = load_workbook(path)
    ws = wb["hoja"]
    header_cell = ws.cell(row=1, column=1)
    assert header_cell.number_format == "@", (
        "La celda de encabezado debe tener number_format '@' para "
        "evitar heredar el formato numerico de la columna"
    )


def test_celda_cliente_id_tiene_formato_texto(tmp_path: Path) -> None:
    """CLIENTE_ID debe tener number_format '@' (texto) a nivel de celda, no formato numerico."""
    df = pd.DataFrame({"NOMBRE_CLIENTE": ["EMPRESA A"], "CLIENTE_ID": [101]})
    path = exportar_excel({"hoja": df}, "test_cliente_id", output_dir=tmp_path)
    wb = load_workbook(path)
    ws = wb["hoja"]
    col_idx = list(df.columns).index("CLIENTE_ID") + 1
    celda = ws.cell(row=2, column=col_idx)
    assert celda.number_format == "@", (
        "CLIENTE_ID debe tener number_format '@' (texto) para no mostrarse como entero"
    )
    assert celda.alignment.horizontal == "left", (
        "CLIENTE_ID debe estar alineado a la izquierda como campo de texto"
    )
