"""Formateador Excel compartido para todos los reportes CxC.

Centraliza constantes de estilo, columnas y funciones de escritura/formato
de hojas Excel con openpyxl, permitiendo que main.py y reporte_cobrador.py
usen el mismo sistema visual sin duplicar código.
"""

from __future__ import annotations

import logging
import tempfile
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from config.settings import SHEET_PASSWORDS

logger = logging.getLogger(__name__)

# ======================================================================
# CONJUNTOS DE COLUMNAS POR TIPO DE FORMATO
# ======================================================================

COLUMNAS_MONEDA: set[str] = {
    "CARGOS", "ABONOS", "IMPORTE", "IMPUESTO",
    "SALDO_FACTURA", "SALDO_CLIENTE",
    "IMPORTE_TOTAL", "IMPORTE_PROMEDIO", "IMPORTE_MAX",
    "TOTAL_CARGOS", "TOTAL_ABONOS", "SALDO",
    "MONTO_CARGO", "MONTO_ABONOS", "DISPONIBLE",
    "SALDO_TOTAL", "SALDO_VIGENTE", "SALDO_VENCIDO",
    "LIMITE_CREDITO", "SALDO_PENDIENTE", "FACTURAS_PAGADAS",
    "FACTURAS_VIGENTES", "IMPUESTO_TOTAL", "MONTO_TOTAL",
    "IMPORTE_AJUSTE", "VENDIDO", "PAGADO",
    "TOTAL_CARGOS_CANCELADOS", "TOTAL_ABONOS_CANCELADOS"
}

_COLUMNAS_MONEDA_PREFIJOS: tuple[str, ...] = ("FACTURAS_VENCIDAS", "VIGENTE:", "VENCIDO:")

COLUMNAS_FECHA_SOLO: set[str] = {
    "FECHA_EMISION",
    "FECHA_VENCIMIENTO",
}

COLUMNAS_FECHA_HORA: set[str] = {
    "FECHA_HORA_CREACION",
    "FECHA_HORA_ULT_MODIF",
    "FECHA_HORA_CANCELACION",
}

COLUMNAS_FECHA: set[str] = COLUMNAS_FECHA_SOLO | COLUMNAS_FECHA_HORA

COLUMNAS_ENTERO: set[str] = {
    "NUM_DOCUMENTOS", "NUM_REGISTROS", "NUM_CARGOS", "NUM_ABONOS",
    "NUM_FACTURAS", "NUM_VENCIDAS", "DIAS_VENCIDO_MAX",
    "NUM_FACTURAS_PENDIENTES", "NUM_FACTURAS_TOTALES",
    "NUM_FACTURAS_VIGENTES", "NUM_FACTURAS_VENCIDAS"
}

# IDs que deben mostrarse como texto para evitar que Excel los interprete como numeros.
COLUMNAS_TEXTO_FORZADO: set[str] = {"CLIENTE_ID"}

COLUMNAS_PORCENTAJE: set[str] = {"PCT_DEL_TOTAL", "UTILIZACION_PCT", "PCT_ACUMULADO", "PCT_VENCIDO", "VALOR"}

COLS_COLOR_CARGOS: set[str] = {"TOTAL_CARGOS", "TOTAL_CARGOS_CANCELADOS", "VENDIDO", "CARGOS"}
COLS_COLOR_ABONOS: set[str] = {"TOTAL_ABONOS", "TOTAL_ABONOS_CANCELADOS", "PAGADO", "ABONOS", "FACTURAS_PAGADAS"}
COLS_COLOR_SALDOS: set[str] = {
    "SALDO_PENDIENTE", "SALDO_VIGENTE", "SALDO_VENCIDO", "SALDO_TOTAL",
    "SALDO", "DISPONIBLE", "LIMITE_CREDITO", "IMPORTE_AJUSTE"
}

PESTANAS_PROTEGIDAS: set[str] = {"registros_totales_cxc"}

COLUMNAS_CALCULADAS_CXC: set[str] = {
    "SALDO_FACTURA", "SALDO_CLIENTE", "DELTA_RECAUDO", "ZSCORE_DELTA_RECAUDO",
    "ATIPICO_DELTA_RECAUDO", "CATEGORIA_RECAUDO", "DELTA_MORA",
    "ZSCORE_DELTA_MORA", "ATIPICO_DELTA_MORA", "CATEGORIA_MORA",
    "ZSCORE_IMPORTE", "ATIPICO_IMPORTE",
}

# ======================================================================
# ESTILOS OPENPYXL
# ======================================================================

_FONT_NAME = "Cambria"
_HEADER_FONT = Font(name=_FONT_NAME, bold=True, color="FFFFFF", size=11)
_FONT_TOTAL  = Font(name=_FONT_NAME, bold=True, size=11)
_FONT_NORMAL = Font(name=_FONT_NAME, size=11)
_FONT_MUTED  = Font(name=_FONT_NAME, color="808080", size=11)

_HEADER_FILL:      PatternFill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
_CALC_HEADER_FILL: PatternFill = PatternFill(start_color="548235", end_color="548235", fill_type="solid")
_BAND_FILL:        PatternFill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
_WHITE_FILL:       PatternFill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
_FILL_TOTAL:       PatternFill = PatternFill(start_color="A6A6A6", end_color="A6A6A6", fill_type="solid")
_FILL_ZERO:        PatternFill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")

FILL_AZUL:     PatternFill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
FILL_VERDE:    PatternFill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
FILL_AMARILLO: PatternFill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
FILL_ROJO:     PatternFill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")

_HEADER_ALIGNMENT: Alignment = Alignment(horizontal="center", vertical="center")
_THIN_BORDER: Border = Border(
    left=Side(style="thin", color="B4C6E7"), right=Side(style="thin", color="B4C6E7"),
    top=Side(style="thin", color="B4C6E7"),  bottom=Side(style="thin", color="B4C6E7"),
)

# ======================================================================
# FUNCIONES DE FORMATO INTERNO
# ======================================================================

def _aplicar_formato_encabezado(ws: Any, n_cols: int, calc_cols: set[str] | None = None) -> None:
    calc_upper: set[str] = {c.upper() for c in calc_cols} if calc_cols else set()
    for col_idx in range(1, n_cols + 1):
        cell = ws.cell(row=1, column=col_idx)
        nombre = str(cell.value).upper() if cell.value else ""
        cell.font = _HEADER_FONT
        cell.fill = _CALC_HEADER_FILL if nombre in calc_upper else _HEADER_FILL
        cell.alignment = _HEADER_ALIGNMENT
        cell.border = _THIN_BORDER
        # Formato texto explicito para evitar que column_dimensions.number_format
        # (asignado por _aplicar_formatos_columna) afecte la fila de cabecera.
        cell.number_format = "@"


def _aplicar_bordes_y_fuente(ws: Any, n_filas: int, n_cols: int) -> None:
    """Aplica borde fino y fuente estandar a todas las celdas de datos."""
    for row_idx in range(2, n_filas + 2):
        for col_idx in range(1, n_cols + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = _THIN_BORDER
            cell.font = _FONT_NORMAL


def _aplicar_formatos_columna(ws: Any, columnas: list[str], n_filas: int, df: pd.DataFrame) -> None:
    """Asigna formatos numericos a nivel de columna (O(n_cols)) en lugar de celda a celda.

    Usa ``ws.column_dimensions[letter].number_format`` para aplicar el formato a
    toda la columna en una sola operacion. Las celdas de encabezado quedan
    protegidas por el formato ``"@"`` explicitamente asignado en
    :func:`_aplicar_formato_encabezado`.

    Caso especial: la columna ``VALOR`` con condicion ``UNIDAD == "%"`` requiere
    evaluacion por fila y se trata con un bucle reducido al final.

    Args:
        ws: Hoja de trabajo openpyxl activa.
        columnas: Lista de nombres de columna en el orden del DataFrame.
        n_filas: Numero de filas de datos (sin contar la cabecera).
        df: DataFrame fuente; solo se usa para el caso especial ``VALOR+UNIDAD``.
    """
    valor_col_idx: int | None = None

    for col_idx, col_name in enumerate(columnas, start=1):
        col_upper = str(col_name).upper()
        letter = get_column_letter(col_idx)
        es_moneda = (
            col_upper in COLUMNAS_MONEDA
            or any(col_upper.startswith(p) for p in _COLUMNAS_MONEDA_PREFIJOS)
        )
        if es_moneda:
            ws.column_dimensions[letter].number_format = "#,##0.00"
        elif col_upper in COLUMNAS_ENTERO:
            ws.column_dimensions[letter].number_format = "0"
        elif col_upper in COLUMNAS_FECHA_SOLO:
            ws.column_dimensions[letter].number_format = "YYYY/MM/DD"
        elif col_upper in COLUMNAS_FECHA_HORA:
            ws.column_dimensions[letter].number_format = "YYYY/MM/DD HH:MM:SS"
        elif col_upper in COLUMNAS_PORCENTAJE:
            if col_upper == "VALOR" and "UNIDAD" in df.columns:
                valor_col_idx = col_idx
            else:
                ws.column_dimensions[letter].number_format = "0.00%"

    # Caso especial: columna VALOR con condicion UNIDAD == "%"
    if valor_col_idx is not None:
        for row_idx in range(2, n_filas + 2):
            unidad = df.iloc[row_idx - 2].get("UNIDAD", "")
            if str(unidad).strip() == "%":
                ws.cell(row=row_idx, column=valor_col_idx).number_format = "0.00%"


def _aplicar_estilos_semanticos(ws: Any, df: pd.DataFrame, columnas: list[str]) -> None:
    for r_idx, row in enumerate(df.itertuples(index=False), start=2):
        row_dict = dict(zip(columnas, row))

        is_total = str(row_dict.get(columnas[0], "")).strip().upper() == "TOTAL"

        is_zero = False
        for saldo_col in ["SALDO_PENDIENTE", "SALDO_TOTAL", "SALDO", "IMPORTE_AJUSTE"]:
            if saldo_col in row_dict:
                val = row_dict[saldo_col]
                if val is not None and str(val).strip() != "":
                    try:
                        if float(val) == 0.0:
                            is_zero = True
                    except (ValueError, TypeError):
                        pass

        clasif = str(row_dict.get("CLASIFICACION", ""))

        for c_idx, col in enumerate(columnas, start=1):
            cell = ws.cell(row=r_idx, column=c_idx)

            if col in ["TOTAL_CARGOS", "TOTAL_CARGOS_CANCELADOS", "LIMITE_CREDITO"]:
                cell.fill = FILL_AZUL
            elif col in ["TOTAL_ABONOS", "TOTAL_ABONOS_CANCELADOS", "SALDO_VIGENTE", "DISPONIBLE", "FACTURAS_PAGADAS"]:
                cell.fill = FILL_VERDE
            elif col in ["SALDO_PENDIENTE", "SALDO_TOTAL", "SALDO", "IMPORTE_AJUSTE"]:
                cell.fill = FILL_AMARILLO
            elif col in ["SALDO_VENCIDO", "DIAS_VENCIDO_MAX", "PCT_VENCIDO"]:
                cell.fill = FILL_ROJO
            else:
                cell.fill = _BAND_FILL if r_idx % 2 == 0 else _WHITE_FILL

            if clasif == "A":
                cell.fill = FILL_VERDE
            elif clasif == "B":
                cell.fill = FILL_AMARILLO
            elif clasif == "C":
                cell.fill = FILL_ROJO

            if is_zero:
                cell.fill = _FILL_ZERO
                cell.font = _FONT_MUTED

            if is_total:
                cell.fill = _FILL_TOTAL
                cell.font = _FONT_TOTAL


_ALIGN_DERECHA:  Alignment = Alignment(horizontal="right",  vertical="center")
_ALIGN_CENTRO:   Alignment = Alignment(horizontal="center", vertical="center")
_ALIGN_IZQUIERDA: Alignment = Alignment(horizontal="left",  vertical="center")


def _aplicar_formato_y_alineacion_datos(
    ws: Any,
    columnas: list[str],
    n_filas: int,
    df: pd.DataFrame,
) -> None:
    """Aplica number_format y alineacion a nivel de CELDA en un unico pass.

    column_dimensions.number_format no aplica a celdas escritas por pandas
    (que tienen un estilo explicito con 'General' que tiene precedencia).
    Esta funcion establece cell.number_format en cada celda de datos para
    garantizar que el formato sea visible en Excel.

    Reglas de alineacion:
        - Moneda, enteros, porcentajes: derecha.
        - Fechas (con o sin hora): centro.
        - Texto, codigos: izquierda.

    Args:
        ws: Hoja de trabajo openpyxl activa.
        columnas: Lista de nombres de columna en el orden del DataFrame.
        n_filas: Numero de filas de datos (sin cabecera).
        df: DataFrame fuente; necesario para el caso especial VALOR+UNIDAD.
    """
    valor_col_idx: int | None = None

    # Pre-calcular (fmt, alignment) por columna para evitar re-evaluacion en el loop de filas
    col_configs: list[tuple[str | None, Alignment]] = []
    for col_name in columnas:
        col_upper = col_name.upper()

        es_moneda = (
            col_upper in COLUMNAS_MONEDA
            or any(col_upper.startswith(p) for p in _COLUMNAS_MONEDA_PREFIJOS)
        )
        es_texto_forzado = col_upper in COLUMNAS_TEXTO_FORZADO
        es_numerico = es_moneda or col_upper in COLUMNAS_ENTERO or col_upper in COLUMNAS_PORCENTAJE
        es_fecha = col_upper in COLUMNAS_FECHA

        if es_texto_forzado or (not es_numerico and not es_fecha):
            align: Alignment = _ALIGN_IZQUIERDA
        elif es_numerico:
            align = _ALIGN_DERECHA
        else:
            align = _ALIGN_CENTRO

        if es_texto_forzado:
            fmt: str | None = "@"
        elif es_moneda:
            fmt = "#,##0.00"
        elif col_upper in COLUMNAS_ENTERO:
            fmt = "0"
        elif col_upper in COLUMNAS_FECHA_SOLO:
            fmt = "YYYY/MM/DD"
        elif col_upper in COLUMNAS_FECHA_HORA:
            fmt = "YYYY/MM/DD HH:MM:SS"
        elif col_upper in COLUMNAS_PORCENTAJE and col_upper != "VALOR":
            fmt = "0.00%"
        else:
            fmt = None

        col_configs.append((fmt, align))
        if col_upper == "VALOR" and "UNIDAD" in df.columns:
            valor_col_idx = len(col_configs) - 1

    for row_idx in range(2, n_filas + 2):
        for col_idx, (fmt, align) in enumerate(col_configs, start=1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.alignment = align
            if fmt is not None:
                cell.number_format = fmt

    # Caso especial: columna VALOR, formato condicional segun UNIDAD
    if valor_col_idx is not None:
        col_1based = valor_col_idx + 1
        for row_idx in range(2, n_filas + 2):
            if str(df.iloc[row_idx - 2].get("UNIDAD", "")).strip() == "%":
                ws.cell(row=row_idx, column=col_1based).number_format = "0.00%"


_ROW_HEIGHT = 15


def _fijar_altura_filas(ws: Any, n_filas: int) -> None:
    """Fija la altura de todas las filas al valor estandar para evitar expansion.

    Args:
        ws: Hoja de trabajo openpyxl activa.
        n_filas: Numero de filas de datos (sin cabecera).
    """
    for row_idx in range(1, n_filas + 2):
        ws.row_dimensions[row_idx].height = _ROW_HEIGHT


def _aplicar_bandas_alternas(ws: Any, band_data: Any, n_cols: int) -> None:
    fill_par = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
    for i, band_value in enumerate(band_data):
        row_idx = i + 2
        fill = fill_par if int(band_value) == 0 else _WHITE_FILL
        for col_idx in range(1, n_cols + 1):
            ws.cell(row=row_idx, column=col_idx).fill = fill


def _autoajustar_ancho_columnas(ws: Any) -> None:
    for col_cells in ws.columns:
        max_length = 0
        col_letter = col_cells[0].column_letter
        header_val = col_cells[0].value

        if header_val:
            max_length = len(str(header_val))

        for cell in col_cells[1:]:
            if cell.value is None or str(cell.value).strip() == "":
                continue
            col_fmt = ws.column_dimensions[col_letter].number_format or ""
            val = cell.value
            if "HH:MM:SS" in col_fmt:
                cell_len = 19
            elif "YYYY/MM/DD" in col_fmt:
                cell_len = 10
            elif "#,##0" in col_fmt or "0.00" in col_fmt:
                try:
                    cell_len = len(f"{float(val):,.2f}")
                except (ValueError, TypeError):
                    cell_len = len(str(val))
            elif "%" in col_fmt:
                try:
                    cell_len = len(f"{float(val)*100:.2f}%")
                except (ValueError, TypeError):
                    cell_len = len(str(val))
            else:
                cell_len = len(str(val))
            if cell_len > max_length:
                max_length = cell_len

        if str(header_val).upper() in ["INTERPRETACION", "MOTIVO"]:
            ws.column_dimensions[col_letter].width = 60
        else:
            adjusted_width = int(max_length * 1.3) + 5
            ws.column_dimensions[col_letter].width = min(max(adjusted_width, 14), 70)


# ======================================================================
# API PÚBLICA
# ======================================================================

def _validar_password_hoja(nombre_hoja: str, password: str | None, protegida: bool) -> None:
    """Lanza ValueError si una hoja protegida no tiene contrasena definida.

    Centraliza el mensaje de error para que tanto ``escribir_hoja`` como
    ``exportar_excel`` (validacion anticipada) produzcan el mismo texto.

    Args:
        nombre_hoja: Nombre de la hoja evaluada.
        password: Valor de la contrasena; puede ser ``None`` si la variable
            de entorno no esta definida.
        protegida: Indica si la hoja esta marcada para proteccion.

    Raises:
        ValueError: Si ``protegida=True`` y ``password`` es ``None`` o vacia.
    """
    if protegida and not password:
        raise ValueError(
            "Se intento proteger la hoja '{}' sin contrasena. "
            "Define EXCEL_SHEET_PASSWORD en el archivo .env.".format(nombre_hoja)
        )


def extraer_banda(df: pd.DataFrame) -> tuple[pd.DataFrame, Any]:
    """Separa la columna _BAND_GROUP del DataFrame y la retorna aparte."""
    if "_BAND_GROUP" in df.columns:
        band_data = df["_BAND_GROUP"].values.copy()
        return df.drop(columns=["_BAND_GROUP"]), band_data
    return df, None


def escribir_hoja(
    writer: Any,
    nombre_hoja: str,
    df: pd.DataFrame,
    band_data: Any = None,
    protegida: bool = False,
    password: str | None = None,
    calc_cols: set[str] | None = None,
) -> None:
    """Escribe un DataFrame en una hoja Excel con todos los estilos corporativos.

    Args:
        writer: Instancia activa de ``pd.ExcelWriter`` con engine openpyxl.
        nombre_hoja: Nombre de la pestana (se trunca a 31 caracteres).
        df: DataFrame a exportar. No debe contener columnas internas como
            ``_BAND_GROUP``.
        band_data: Arreglo de valores 0/1 para bandas alternas por grupo de
            factura. Si es ``None`` se aplican estilos semanticos por fila.
        protegida: Si es ``True``, activa la proteccion de hoja en Excel.
            Requiere que ``password`` sea una cadena no vacia.
        password: Contrasena para la proteccion de hoja. Obligatoria cuando
            ``protegida=True``; se ignora en caso contrario.
        calc_cols: Conjunto de nombres de columnas calculadas que reciben un
            color de encabezado diferente.

    Raises:
        ValueError: Si ``protegida=True`` y ``password`` es ``None`` o vacia.
            Indica que ``EXCEL_SHEET_PASSWORD`` no esta definida en ``.env``.
    """
    _validar_password_hoja(nombre_hoja, password, protegida)
    sheet_name = nombre_hoja[:31]
    df.to_excel(writer, sheet_name=sheet_name, index=False)
    ws = writer.sheets[sheet_name]

    n_filas = len(df)
    n_cols  = len(df.columns)
    columnas = [str(c) for c in df.columns]

    _aplicar_formato_encabezado(ws, n_cols, calc_cols=calc_cols)
    _aplicar_bordes_y_fuente(ws, n_filas, n_cols)
    _aplicar_formatos_columna(ws, columnas, n_filas, df)
    _aplicar_formato_y_alineacion_datos(ws, columnas, n_filas, df)
    _fijar_altura_filas(ws, n_filas)

    if band_data is not None:
        _aplicar_bandas_alternas(ws, band_data, n_cols)
    else:
        _aplicar_estilos_semanticos(ws, df, columnas)

    _autoajustar_ancho_columnas(ws)
    ws.sheet_view.showGridLines = False

    if protegida:
        ws.protection.sheet = True
        ws.protection.password = password

    logger.info("  Hoja '%s': %d filas%s", sheet_name, n_filas, " (protegida)" if protegida else "")


def exportar_excel(
    dataframes: dict[str, pd.DataFrame],
    nombre_base: str,
    output_dir: Path | None = None,
    timestamp: str = "",
    orden_hojas: list[str] | None = None,
    cols_calc_por_hoja: dict[str, set[str]] | None = None,
) -> Path:
    """Exporta un diccionario de DataFrames a un archivo Excel con estilos corporativos.

    Returns:
        Ruta del archivo .xlsx generado.
    """
    if orden_hojas is None:
        orden_hojas = list(dataframes.keys())
    if cols_calc_por_hoja is None:
        cols_calc_por_hoja = {}

    # Validacion anticipada: verificar passwords antes de abrir el writer.
    # Evita que openpyxl lance IndexError al intentar salvar un workbook
    # vacio durante el cleanup de una excepcion interna.
    todas_hojas = [*orden_hojas, *[k for k in dataframes if k not in orden_hojas]]
    for _hoja in todas_hojas:
        if _hoja in dataframes:
            _df = dataframes[_hoja]
            if _df is not None and not getattr(_df, "empty", True):
                _validar_password_hoja(
                    _hoja, SHEET_PASSWORDS.get(_hoja), _hoja in PESTANAS_PROTEGIDAS
                )

    sufijo = f"_{timestamp}" if timestamp else ""
    nombre_archivo = f"{nombre_base}{sufijo}".upper() + ".xlsx"

    if output_dir:
        filepath = Path(output_dir) / nombre_archivo
        filepath.parent.mkdir(parents=True, exist_ok=True)
    else:
        logger.warning("No se detecto output_dir. Usando directorio temporal.")
        filepath = Path(tempfile.gettempdir()) / nombre_archivo

    with pd.ExcelWriter(filepath, engine="openpyxl") as writer:
        hojas_escritas = 0

        for nombre_hoja in orden_hojas:
            if nombre_hoja in dataframes:
                df = dataframes[nombre_hoja]
                if df is not None and not df.empty:
                    df_limpio, band_data = extraer_banda(df)
                    escribir_hoja(
                        writer, nombre_hoja, df_limpio,
                        band_data=band_data,
                        protegida=nombre_hoja in PESTANAS_PROTEGIDAS,
                        password=SHEET_PASSWORDS.get(nombre_hoja),
                        calc_cols=cols_calc_por_hoja.get(nombre_hoja),
                    )
                    hojas_escritas += 1

        for nombre_hoja, df in dataframes.items():
            if nombre_hoja not in orden_hojas and df is not None and not df.empty:
                df_limpio, band_data = extraer_banda(df)
                escribir_hoja(
                    writer, nombre_hoja, df_limpio,
                    band_data=band_data,
                    protegida=nombre_hoja in PESTANAS_PROTEGIDAS,
                    password=SHEET_PASSWORDS.get(nombre_hoja),
                    calc_cols=cols_calc_por_hoja.get(nombre_hoja),
                )
                hojas_escritas += 1

        if hojas_escritas == 0:
            logger.warning("Ninguna hoja activa para %s. Inyectando contingencia.", filepath.name)
            pd.DataFrame({"AVISO": ["Ausencia de datos transaccionales en este periodo"]}).to_excel(
                writer, sheet_name="Sin Datos", index=False
            )

    return filepath
